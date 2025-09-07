import sys
import openpyxl as excel
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill
import pandas as pd
import os

from PyQt5.QtWidgets import QApplication, QWidget, QPushButton, QHBoxLayout, QVBoxLayout, QComboBox, QLabel
from PyQt5.QtCore import QCoreApplication

if getattr(sys, 'frozen', False):
    # test.exe로 실행한 경우,test.exe를 보관한 디렉토리의 full path를 취득
    program_directory = os.path.dirname(os.path.abspath(sys.executable))
else:
    # python test.py로 실행한 경우,test.py를 보관한 디렉토리의 full path를 취득
    program_directory = os.path.dirname(os.path.abspath(__file__))

class MyApp(QWidget):

    def __init__(self):
        super().__init__()
        self.wb = excel.Workbook()
        self.wb.active.title = 'label'
        self.wb.create_sheet('template1', 1)
        self.w1 = self.wb['label']
        self.w2 = self.wb['template1']

        # writing
        self.style_center = Alignment(horizontal='center', vertical='center')
        self.title_font = Font(size=20, bold=True, color='ffffff')
        self.tile_font = Font(size=15, bold=True, color='000000')
        self.subject_bg = PatternFill('solid', fgColor='4BACC6')
        self.AOC_bg = PatternFill('solid', fgColor='ff0000')
        self.copper_bg = PatternFill('solid', fgColor='0000FF')
        self.s_tile_font = Font(size=10, bold=True, color='000000')
        self.b_tile_font = Font(size=20, bold=True, color='000000')

        self.w1.merge_cells('A1:G1')
        self.w1['A1'] = "Label print maker"
        self.w1['A1'].font = self.title_font
        self.w1['A1'].fill = self.subject_bg
        self.w1['A1'].alignment = self.style_center

        self.w2.merge_cells('A1:G1')
        self.w2['A1'] = "Label template"
        self.w2['A1'].font = self.title_font
        self.w2['A1'].fill = self.subject_bg
        self.w2['A1'].alignment = self.style_center

        self.initUI()

    def initUI(self):

        PRD_Button = QPushButton('PRD make!', self)
        MOR_Button = QPushButton('MOR make!', self)
        cancelButton = QPushButton('Cancel', self)

        self.name = QLabel('Option1', self)
        self.name.move(50, 150)

        file_list = os.listdir()
        cb = QComboBox(self)
        for item in file_list:
            cb.addItem(item)
        cb.move(50, 50)

        cb.activated[str].connect(self.setting_name)

        PRD_Button.clicked.connect(self.PRD)
        MOR_Button.clicked.connect(self.MOR)
        cancelButton.clicked.connect(QCoreApplication.instance().quit)

        hbox = QHBoxLayout()
        hbox.addStretch(1)
        hbox.addWidget(cancelButton)
        hbox.addStretch(1)

        hbox1 = QHBoxLayout()
        hbox1.addStretch(1)
        hbox1.addWidget(cb)
        hbox1.addWidget(self.name)
        hbox1.addStretch(1)

        hbox2 = QHBoxLayout()
        hbox2.addStretch(1)
        hbox2.addWidget(PRD_Button)
        hbox2.addWidget(MOR_Button)
        hbox2.addStretch(1)

        vbox = QVBoxLayout()
        vbox.addStretch(1)
        vbox.addLayout(hbox1)
        vbox.addStretch(1)
        vbox.addLayout(hbox2)
        vbox.addStretch(1)
        vbox.addLayout(hbox)
        vbox.addStretch(1)

        self.setLayout(vbox)

        self.setWindowTitle('Box Layout')
        self.setGeometry(300, 300, 300, 200)
        self.show()

    def setting_name(self, text):
        self.name.setText(text)

    def PRD(self):
        default = "Final_Result"
        load_file_name = self.name.text()
        data = pd.read_excel(load_file_name, sheet_name=default, na_values='=')
        max_length = 0
        dump = 4
        rack_dump = []

        # Observe all Start point
        AOC_dict_start, total_AOC_s = dict(), []
        AOC_dict_end, total_AOC_e = dict(), []
        COP_dict_start, total_COP_s = dict(), []
        COP_dict_end, total_COP_e = dict(), []
        for idx, value in enumerate(data.values):
            # print(idx, " : ", value[3][12:15])
            AOC_dict_start[value[3][:5]] = []
            AOC_dict_end[value[3][:5]] = []
            COP_dict_start[value[3][:5]] = []
            COP_dict_end[value[3][:5]] = []

        # insert table
        for idx, value in enumerate(data.values):
            if value[-2] == 100000:
                AOC_dict_start[value[3][:5]].append(value[3])
                AOC_dict_end[value[3][:5]].append(value[4])
            else:
                COP_dict_start[value[3][:5]].append(value[3])
                COP_dict_end[value[3][:5]].append(value[4])

        Key = list(AOC_dict_start.keys())
        temp_1, temp_2, temp_3, temp_4, temp_reack = [], [], [], [], []
        # write in w1
        for id, text in enumerate(Key):
            self.w1.cell(7 * id + 2, 1).value = text
            self.w1.cell(7 * id + 2, 1).font = self.tile_font
            self.w1.cell(7 * id + 3, 1).value = "AOC cable"
            self.w1.cell(7 * id + 3, 1).fill = self.AOC_bg
            self.w1.append(AOC_dict_start[text] * 2)
            self.w1.append(AOC_dict_end[text] * 2)
            self.w1.cell(7 * id + 6, 1).value = "Copper cable"
            self.w1.cell(7 * id + 6, 1).fill = self.copper_bg
            self.w1.append(COP_dict_start[text] * 2)
            self.w1.append(COP_dict_end[text] * 2)

            temp_1 += AOC_dict_start[text] * 2
            temp_2 += AOC_dict_end[text] * 2
            temp_3 += COP_dict_start[text] * 2
            temp_4 += COP_dict_end[text] * 2
            temp_reack.append(text)

            if id % dump == dump-1 or (id // dump == len(Key) // dump and id % dump == len(Key) % dump):
                total_AOC_s.append(temp_1)
                total_AOC_e.append(temp_2)
                total_COP_s.append(temp_3)
                total_COP_e.append(temp_4)
                rack_dump.append(temp_reack)
                temp_1, temp_2, temp_3, temp_4, temp_reack = [], [], [], [], []


            if max_length < len(AOC_dict_start[text] * 2):
                max_length = len(AOC_dict_start[text] * 2)

        for id, label in enumerate(rack_dump):
            self.w2.cell(9 * id + 2, 2).value = "Cop_label : " + str(len(total_COP_s[id])) + "EA"
            self.w2.cell(9 * id + 2, 1).value = "AOC_label : " + str(len(total_AOC_s[id])) + "EA"
            self.w2.append(label)
            self.w2.cell(9 * id + 4, 1).value = "AOC cable"
            self.w2.cell(9 * id + 4, 1).fill = self.AOC_bg
            self.w2.append(total_AOC_s[id])
            self.w2.append(total_AOC_e[id])
            self.w2.cell(9 * id + 7, 1).value = "Copper cable"
            self.w2.cell(9 * id + 7, 1).fill = self.copper_bg
            self.w2.append(total_COP_s[id])
            self.w2.append(total_COP_e[id])
            self.w2.append([])

        for col in range(1, max_length + 1):
            self.w1.column_dimensions[get_column_letter(col)].width = 22
            self.w2.column_dimensions[get_column_letter(col)].width = 22

        for col in range(1, max_length * dump + 1):
            self.w2.column_dimensions[get_column_letter(col)].width = 22
            self.w2.column_dimensions[get_column_letter(col)].width = 22

        self.wb.save("label_filter.xlsx")
        QCoreApplication.instance().quit()

    def MOR(self):
        default = "NDT"
        load_file_name = self.name.text()
        data = pd.read_excel(load_file_name, sheet_name=default)
        idx_device, idx_SPort, idx_Start, idx_End, idx_EndDev = 0, 1, 4, 9, 11
        Copper = ['Gi0/0/1', 'Gi0/0/2', 'CONS', 'MGMT', 'COM6', 'NIC1']
        Cooper_dict = {'Long_Start': [], 'Long_End': [], 'Short_Start': [], 'Short_End': []}
        LC_dict = {'Long_Start': [], 'Long_End': [], 'Short_Start': [], 'Short_End': []}
        MPO = {'Start': [], 'End': []}
        PSM4_dict_2m = {'Start': [], 'End': []}
        PSM4_dict = {}
        T1_num = 8
        for value in range(T1_num):
            PSM4_dict[str(value) + 'Start'] = []
            PSM4_dict[str(value) + 'End'] = []
        T1_flag = 0

        for device in data.values:
            if device[idx_SPort] in Copper:
                idx = idx_Start
                while idx <= idx_End:
                    if pd.isnull(device[idx]):
                        pass
                    else:
                        if device[idx_device][-2:] in ['P1', 'P2', 'T1']:
                            if (idx - idx_Start) % 2 == 0:
                                Cooper_dict['Long_Start'].append(device[idx])
                            else:
                                Cooper_dict['Long_End'].append(device[idx])
                        else:
                            if (idx - idx_Start) % 2 == 0:
                                Cooper_dict['Short_Start'].append(device[idx])
                            else:
                                Cooper_dict['Short_End'].append(device[idx])
                    idx += 1
            if device[idx_SPort] == 'Gi0/0/0':
                idx = idx_Start
                while idx <= idx_End:
                    if pd.isnull(device[idx]):
                        pass
                    else:
                        if (idx - idx_Start) // 2 == 0:
                            if (idx - idx_Start) % 2 == 0:
                                LC_dict['Short_Start'].append(device[idx])
                            else:
                                LC_dict['Short_End'].append(device[idx])
                        else:
                            if (idx - idx_Start) % 2 == 0:
                                LC_dict['Long_Start'].append(device[idx])
                            else:
                                LC_dict['Long_End'].append(device[idx])
                    idx += 1
            if device[idx_device][-2:] == 'M0' and device[idx_SPort][:3] == 'ETH':
                idx = idx_Start
                while idx <= idx_End:
                    if (idx - idx_Start) // 2 == 0:
                        if (idx - idx_Start) % 2 == 0:
                            LC_dict['Short_Start'].append(device[idx])
                        else:
                            LC_dict['Short_End'].append(device[idx])
                    elif (idx - idx_Start) // 2 == 1:
                        if (idx - idx_Start) % 2 == 0:
                            LC_dict['Long_Start'].append(device[idx])
                        else:
                            LC_dict['Long_End'].append(device[idx])
                    else:
                        if (idx - idx_Start) % 2 == 0:
                            MPO['Start'].append(device[idx])
                        else:
                            MPO['End'].append(device[idx])
                    idx += 1
            if device[idx_device][-2:] == 'T1' and device[idx_SPort][:3] == 'ETH':
                idx = idx_Start
                while idx <= idx_End:
                    if pd.isnull(device[idx]):
                        pass
                    else:
                        if (idx - idx_Start) // 2 == 0:
                            if (idx - idx_Start) % 2 == 0:
                                PSM4_dict_2m['Start'].append(device[idx])
                            else:
                                PSM4_dict_2m['End'].append(device[idx])
                        else:
                            id = (int(device[idx_EndDev][-4:-2]) - 1) % 8
                            if (idx - idx_Start) % 2 == 0:
                                PSM4_dict[str(id) + 'Start'].append(device[idx])
                            else:
                                PSM4_dict[str(id) + 'End'].append(device[idx])
                    idx += 1

        Cooper_dict['Short_Start'].pop(0)
        Cooper_dict['Short_End'].pop(0)

        # MOR SIDE
        self.w1.cell(2, 1).value = "MOR SIDE"
        self.w1.cell(2, 1).font = self.b_tile_font

        for name in Cooper_dict.keys():
            self.w1.append(Cooper_dict[name] * 2)

        self.w1.insert_rows(3)
        self.w1.cell(3, 1).value = "Long Copper [12m]"
        self.w1.cell(3, 2).value = str(len(Cooper_dict['Long_Start'])) + "EA"
        self.w1.cell(3, 1).font = self.tile_font

        self.w1.insert_rows(6)
        self.w1.cell(6, 1).value = "Short Copper[7m]"
        self.w1.cell(6, 2).value = str(len(Cooper_dict['Short_Start'])) + "EA"
        self.w1.cell(6, 1).font = self.tile_font

        self.w1.cell(9, 1).value = "SM-LC (2m)"
        self.w1.cell(9, 2).value = str(len(LC_dict['Short_Start'])) + "EA"
        self.w1.cell(9, 1).font = self.tile_font

        for name in LC_dict.keys():
            if "Short" in name:
                self.w1.append(LC_dict[name] * 2)

        self.w1.cell(12, 1).value = "PSM4 (2m)"
        self.w1.cell(12, 2).value = str(len(PSM4_dict_2m['Start'])) + "EA"
        self.w1.cell(12, 1).font = self.tile_font

        for i in range(T1_num):
            self.w1.cell(13 + (3 * i), 1).value = PSM4_dict_2m['Start'][8 * i][:11]
            self.w1.cell(13 + (3 * i), 1).font = self.s_tile_font
            self.w1.append(PSM4_dict_2m['Start'][8 * i:8 * (i + 1)] * 2)
            self.w1.append(PSM4_dict_2m['End'][8 * i:8 * (i + 1)] * 2)

        # MOR SIDE
        self.w1.cell(40, 1).value = "IDF SIDE"
        self.w1.cell(40, 1).font = self.b_tile_font

        self.w1.cell(41, 1).value = "SM-LC (9m)"
        self.w1.cell(41, 2).value = str(len(LC_dict['Long_Start'])) + "EA"
        self.w1.cell(41, 1).font = self.tile_font

        for name in LC_dict.keys():
            if "Long" in name:
                self.w1.append(LC_dict[name] * 2)

        self.w1.cell(44, 1).value = "PSM4 "
        self.w1.cell(44, 2).value = "Each T2 " + str(len(PSM4_dict['0Start'])) + "EA"
        self.w1.cell(44, 1).font = self.tile_font

        for i in range(T1_num):
            self.w1.cell(45 + (3 * i), 1).value = str(i + 1) + "&" + str(i + 9) + "T2"
            self.w1.cell(45 + (3 * i), 1).font = self.s_tile_font
            self.w1.append(PSM4_dict[str(i) + 'Start'] * 2)
            self.w1.append(PSM4_dict[str(i) + 'End'] * 2)

        for col in range(1, len(PSM4_dict_2m['Start']) + 1):
            self.w1.column_dimensions[get_column_letter(col)].width = 24
            self.w2.column_dimensions[get_column_letter(col)].width = 24

        self.wb.save("label_filter.xlsx")
        QCoreApplication.instance().quit()

app = QApplication(sys.argv)
ex = MyApp()
sys.exit(app.exec_())