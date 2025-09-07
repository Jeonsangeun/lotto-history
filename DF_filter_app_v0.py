import sys
from openpyxl.utils import get_column_letter
from collections import deque
from openpyxl import load_workbook

from PyQt5.QtWidgets import QApplication, QWidget, QPushButton, QHBoxLayout, QVBoxLayout, QMessageBox, QLabel, QFileDialog


def find_logic(text):
    flag_str = deque(maxlen=3)
    device_length = 20 # max 22 '-' : 4
    save_trigger, end_device, end_port, ban_save = 0, 0, 0, 0
    Device_dic = []
    Device, port = 'PU', ''
    dev_symbol = [' ', ':']
    prt_symbol = [' ', '<', ':', ';', '\n']

    for iid, word in enumerate(text):
        flag_str.append(word)
        txt = ''.join(flag_str)
        # Find PUS device
        if txt == 'PUS' or txt == 'pus':
            save_trigger = 1
        # Device & Port edge point setting
        if save_trigger == 1:
            if end_device == 1 and (word in prt_symbol):
                end_port = 1
            if word in dev_symbol:
                end_device = 1
        # Extract each case
        if save_trigger == 1 and end_device == 0 and end_port == 0:
            if word == 's':
                word = 'S'
            Device += word
        elif save_trigger == 1 and end_device == 1 and end_port == 0:
            if len(Device) == device_length or len(Device) == device_length + 1:
                port += word
            else:
                ban_save = 1
        # Finishing detect device&port
        if end_port == 1 or iid == len(text)-1:
            if ban_save == 0 and len(port) > 2:
                Device_dic.append((Device, port[1:]))
            save_trigger, end_device, end_port, ban_save, Device, port = 0, 0, 0, 0, 'PU', ''

    return Device_dic

# From reference
def AutoFitColumnSize(worksheet, columns=None, rows=None, margin=2):
    for i, column_cells in enumerate(worksheet.columns):
        is_ok = False
        if columns == None:
            is_ok = True
        elif isinstance(columns, list) and i in columns:
            is_ok = True

        if is_ok:
            length = max(len(str(cell.value)) for cell in column_cells)
            worksheet.column_dimensions[column_cells[0].column_letter].width = length + margin

    return worksheet

class MyApp(QWidget):

    def __init__(self):
        super().__init__()
        self.url = QLabel('url', self)

        self.initUI()

    def initUI(self):
        F_Button = QPushButton('Search', self)
        F_Button.clicked.connect(self.search_button_clicked)

        PRD_Button = QPushButton('Go!', self)
        PRD_Button.clicked.connect(self.PRD)

        hbox = QHBoxLayout()
        hbox.addStretch(1)
        hbox.addWidget(F_Button)
        hbox.addStretch(1)
        hbox.addWidget(self.url)
        hbox.addStretch(1)

        hbox2 = QHBoxLayout()
        hbox2.addWidget(PRD_Button)

        vbox = QVBoxLayout()
        vbox.addStretch(1)
        vbox.addLayout(hbox)
        vbox.addStretch(1)
        vbox.addLayout(hbox2)
        vbox.addStretch(1)

        self.setLayout(vbox)

        self.setWindowTitle('Box Layout')
        self.setGeometry(300, 300, 300, 200)
        self.show()

    def search_button_clicked(self):
        f_path = QFileDialog.getOpenFileName(self, 'Open File', '')
        self.url.setText(f_path[0])

    def PRD(self):
        load_file_name = self.url.text()
        wb = load_workbook(load_file_name, data_only=True)
        ws = wb.active  ## 첫 번째 시트
        row_title = ws[1]
        title_list = []
        Insert_list = ['End Port', 'End Device', 'Start Port', 'Start Device']
        Device_data = []
        for title in row_title:
            title_list.append(title.value)

        Des_index = title_list.index('Description') + 2
        max_row = ws.max_row

        for col in range(2, max_row + 1):
            temp = find_logic(str(ws.cell(col, Des_index-1).value))
            if temp == []:
                temp = find_logic(str(ws.cell(col, Des_index-2).value))
            Device_data.append(temp)

        # [Description Start Device Start Port End Device End Port]
        for plus_title in Insert_list:
            ws.insert_cols(Des_index)
            ws.cell(1, Des_index).value = plus_title

        Device_data.reverse()
        start_row = 2

        while len(Device_data) > 0:
            Pair_list, resist_list = [], [] # resist_lit : [[Device, port], [Device, Port]]
            for idx, data in enumerate(Device_data[-1]):
                resist_list.append(data)
                if idx % 2 == 1:
                    Pair_list.append(resist_list)
                    resist_list = []

            for data_set in Pair_list:
                if data_set[0][0][:5] == 'PUS04':
                    ws.cell(start_row, Des_index).value = data_set[0][0]  # Start device
                    ws.cell(start_row, Des_index + 1).value = data_set[0][1]  # Start port
                    ws.cell(start_row, Des_index + 2).value = data_set[1][0]  # end device
                    ws.cell(start_row, Des_index + 3).value =data_set[1][1]  # end port
                else:
                    ws.cell(start_row, Des_index).value = data_set[1][0]  # Start device
                    ws.cell(start_row, Des_index + 1).value = data_set[1][1]  # Start port
                    ws.cell(start_row, Des_index + 2).value = data_set[0][0]  # end device
                    ws.cell(start_row, Des_index + 3).value = data_set[0][1]  # end port
                start_row += 1
                if Pair_list[-1] != data_set:
                    ws.insert_rows(start_row)

            Device_data.pop()

        ws = AutoFitColumnSize(ws)
        ws.column_dimensions[get_column_letter(Des_index-1)].width = 22

        file_name = 'Save_to_DF.xlsx'
        wb.save(file_name)
        wb.close()  ## Workbook 종료
        QMessageBox.about(self, 'Finish', 'File name is Save_to_DF.xlsx')

app = QApplication(sys.argv)
ex = MyApp()
sys.exit(app.exec_())