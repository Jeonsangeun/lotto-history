import sys
import openpyxl as excel
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill
import pandas as pd

from PyQt5.QtWidgets import QApplication, QWidget, QPushButton, QHBoxLayout, QVBoxLayout, QComboBox, QLabel, \
    QFileDialog, QMessageBox
from PyQt5.QtCore import QCoreApplication

class MyApp(QWidget):

    def __init__(self):
        super().__init__()
        self.wb = excel.Workbook()
        self.wb.active.title = 'label'
        self.wb.create_sheet('template1', 1)
        self.w1 = self.wb['label']
        self.w2 = self.wb['template1']

        # writing
        self.style_center = Alignment(horizontal='center', vertical='center')
        self.title_font = Font(size=20, bold=True, color='ffffff')
        self.tile_font = Font(size=15, bold=True, color='000000')
        self.subject_bg = PatternFill('solid', fgColor='4BACC6')
        self.AOC_bg = PatternFill('solid', fgColor='ff0000')
        self.copper_bg = PatternFill('solid', fgColor='0000FF')
        self.s_tile_font = Font(size=10, bold=True, color='000000')
        self.b_tile_font = Font(size=20, bold=True, color='000000')

        self.w1.merge_cells('A1:G1')
        self.w1['A1'] = "Label print maker"
        self.w1['A1'].font = self.title_font
        self.w1['A1'].fill = self.subject_bg
        self.w1['A1'].alignment = self.style_center

        self.w2.merge_cells('A1:G1')
        self.w2['A1'] = "Label template"
        self.w2['A1'].font = self.title_font
        self.w2['A1'].fill = self.subject_bg
        self.w2['A1'].alignment = self.style_center

        self.initUI()

    def initUI(self):

        PRD_Button = QPushButton('PRD make!', self)
        MOR_Button = QPushButton('MOR make!', self)
        cancelButton = QPushButton('Cancel', self)

        self.name = QLabel('Option1', self)
        self.name.move(50, 150)

        # file_list = os.listdir()
        # cb = QComboBox(self)
        # for item in file_list:
        #     cb.addItem(item)
        # cb.move(50, 50)
        # cb.activated[str].connect(self.setting_name)

        res_url = QPushButton("NDT URL")
        res_url.clicked.connect(self.url_clicked)

        PRD_Button.clicked.connect(self.PRD)
        MOR_Button.clicked.connect(self.MOR)
        cancelButton.clicked.connect(QCoreApplication.instance().quit)

        hbox = QHBoxLayout()
        hbox.addStretch(1)
        hbox.addWidget(cancelButton)
        hbox.addStretch(1)

        hbox1 = QHBoxLayout()
        hbox1.addStretch(1)
        hbox1.addWidget(res_url)
        hbox1.addWidget(self.name)
        hbox1.addStretch(1)

        hbox2 = QHBoxLayout()
        hbox2.addStretch(1)
        hbox2.addWidget(PRD_Button)
        hbox2.addWidget(MOR_Button)
        hbox2.addStretch(1)

        vbox = QVBoxLayout()
        vbox.addStretch(1)
        vbox.addLayout(hbox1)
        vbox.addStretch(1)
        vbox.addLayout(hbox2)
        vbox.addStretch(1)
        vbox.addLayout(hbox)
        vbox.addStretch(1)

        self.setLayout(vbox)

        self.setWindowTitle('Box Layout')
        self.setGeometry(300, 300, 300, 200)
        self.show()

    def setting_name(self, text):
        self.name.setText(text)

    def url_clicked(self):
        f_path = QFileDialog.getOpenFileName(self, 'Open File', '', 'EXCEL(*.xlsx)')
        if f_path[0][-5:] != '.xlsx':
            QMessageBox.about(self, 'Alert', 'Please select .xlsx')
        else:
            self.name.setText(f_path[0])

    def PRD(self):
        default = "Final_Result"
        load_file_name = self.name.text()
        data = pd.read_excel(load_file_name, sheet_name=default, na_values='=')
        PRD_height = 9
        Fast_PRD_height = 11
        dump = 4
        rack_dump = []

        # indexing
        index_STL = list(data.keys()).index('Start Label')  # Start Label
        index_EDL = list(data.keys()).index('End Label')  # Start Label
        index_SPD = list(data.keys()).index('Speed')  # Start Label

        # Observe all Start point
        AOC_dict_start, total_AOC_s = dict(), []
        AOC_dict_end, total_AOC_e = dict(), []
        COP_dict_start, total_COP_s = dict(), []
        COP_dict_end, total_COP_e = dict(), []

        # Direct label format
        total_AOC, total_COP = [], []

        for idx, value in enumerate(data.values):
            # print(idx, " : ", value[3][12:15])
            AOC_dict_start[value[index_STL][:5]] = []
            AOC_dict_end[value[index_STL][:5]] = []
            COP_dict_start[value[index_STL][:5]] = []
            COP_dict_end[value[index_STL][:5]] = []

        # insert table
        for idx, value in enumerate(data.values):
            if value[index_SPD] == 100000:  # speed
                AOC_dict_start[value[index_STL][:5]].append(value[index_STL])
                AOC_dict_end[value[index_STL][:5]].append(value[index_EDL])
            else:
                COP_dict_start[value[index_STL][:5]].append(value[index_STL])
                COP_dict_end[value[index_STL][:5]].append(value[index_EDL])

        Key = list(AOC_dict_start.keys())
        temp_1, temp_2, temp_3, temp_4, temp_5, temp_6, temp_rack = [], [], [], [], [], [], []
        # write in w1
        for id, text in enumerate(Key):
            temp1 = [i + "\n" + j for i, j in zip(AOC_dict_start[text], AOC_dict_end[text])]
            temp2 = [i + "\n" + j for i, j in zip(COP_dict_start[text], COP_dict_end[text])]
            self.w1.cell(PRD_height * id + 2, 1).value = text
            self.w1.cell(PRD_height * id + 2, 1).font = self.tile_font
            self.w1.cell(PRD_height * id + 3, 1).value = "AOC cable"
            self.w1.cell(PRD_height * id + 3, 1).fill = self.AOC_bg
            self.w1.append(AOC_dict_start[text] * 2)
            self.w1.append(AOC_dict_end[text] * 2)
            self.w1.append(temp1 * 2)
            self.w1.cell(PRD_height * id + 7, 1).value = "Copper cable"
            self.w1.cell(PRD_height * id + 7, 1).fill = self.copper_bg
            self.w1.append(COP_dict_start[text] * 2)
            self.w1.append(COP_dict_end[text] * 2)
            self.w1.append(temp2 * 2)

            temp_1 += AOC_dict_start[text] * 2
            temp_2 += AOC_dict_end[text] * 2
            temp_3 += COP_dict_start[text] * 2
            temp_4 += COP_dict_end[text] * 2
            temp_5 += temp1 * 2
            temp_6 += temp2 * 2
            temp_rack.append(text)

            if id % dump == dump - 1 or id == len(Key) - 1:
                total_AOC_s.append(temp_1)
                total_AOC_e.append(temp_2)
                total_COP_s.append(temp_3)
                total_COP_e.append(temp_4)
                total_AOC.append(temp_5)
                total_COP.append(temp_6)
                rack_dump.append(temp_rack)
                temp_1, temp_2, temp_3, temp_4, temp_5, temp_6, temp_rack = [], [], [], [], [], [], []

        for id, label in enumerate(rack_dump):
            self.w2.cell(Fast_PRD_height * id + 2, 2).value = "Cop_label : " + str(len(total_COP_s[id])) + "EA"
            self.w2.cell(Fast_PRD_height * id + 2, 1).value = "AOC_label : " + str(len(total_AOC_s[id])) + "EA"
            self.w2.append(label)
            self.w2.cell(Fast_PRD_height * id + 4, 1).value = "AOC cable"
            self.w2.cell(Fast_PRD_height * id + 4, 1).fill = self.AOC_bg
            self.w2.append(total_AOC_s[id])
            self.w2.append(total_AOC_e[id])
            self.w2.append(total_AOC[id])
            self.w2.cell(Fast_PRD_height * id + 8, 1).value = "Copper cable"
            self.w2.cell(Fast_PRD_height * id + 8, 1).fill = self.copper_bg
            self.w2.append(total_COP_s[id])
            self.w2.append(total_COP_e[id])
            self.w2.append(total_COP[id])
            self.w2.append([])

        # 열, 행 너비 맞춤
        for col in range(1, self.w1.max_column + 1):
            self.w1.column_dimensions[get_column_letter(col)].width = 22

        for row in range(len(Key)):
            self.w1.row_dimensions[row * PRD_height + 6].height = 30
            for cell1 in self.w1[row * PRD_height + 6]:
                cell1.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell1.font = Font(size=5.5)
            self.w1.row_dimensions[row * PRD_height + 10].height = 30
            for cell2 in self.w1[row * PRD_height + 10]:
                cell2.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell2.font = Font(size=5.5)

        for col in range(1, self.w2.max_column + 1):
            self.w2.column_dimensions[get_column_letter(col)].width = 22

        for row in range(len(rack_dump)):
            self.w2.row_dimensions[row * Fast_PRD_height + 7].height = 30
            for cell1 in self.w2[row * Fast_PRD_height + 7]:
                cell1.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell1.font = Font(size=5.5)
            self.w2.row_dimensions[row * Fast_PRD_height + 11].height = 30
            for cell2 in self.w2[row * Fast_PRD_height + 11]:
                cell2.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell2.font = Font(size=5.5)

        self.wb.save(load_file_name+"PRD_label_filter.xlsx")
        QMessageBox.about(self, 'Alert', 'Completed!')
        QCoreApplication.instance().quit()

    def MOR(self):
        default = "NDT"
        load_file_name = self.name.text()
        data = pd.read_excel(load_file_name, sheet_name=default)
        # index for title colum
        Key_title = list(data.keys())
        # idx_device, idx_SPort, idx_Start, idx_End, idx_EndDev = 0, 1, 4, 9, 11
        # StartPort = list(filter(lambda x: Key_title[x] == 'StartPort', range(len(Key_title)))) 중복일 경우

        idx_device = Key_title.index('#Fields:StartDevice')  # #Fields:StartDevice
        idx_SPort = Key_title.index('StartPort')  # StartPort1
        idx_Start = Key_title.index('StartPort.1')  # StartPort2
        idx_End = Key_title.index('EndPort')  # EndPort
        idx_EndDev = Key_title.index('EndDevice')  # EndDevice
        idx_LinkType = Key_title.index('LinkType')

        # print(idx_device, idx_SPort, idx_Start, idx_End, idx_EndDev)

        # Make dict initialized
        Copper = ['Gi0/0/1', 'Gi0/0/2']
        Cooper_dict = {'Long_Start': [], 'Long_End': [], 'Short_Start': [], 'Short_End': []}
        LC_dict = {'Long_Start': [], 'Long_End': [], 'Short_Start': [], 'Short_End': []}
        MPO = {'Start': [], 'End': []}
        PSM4_dict_2m = {'Start': [], 'End': []}
        PSM4_dict = {}
        T1_num = 8

        PSM4_list = []
        for value in range(T1_num):
            PSM4_dict[str(value) + 'Start'] = []
            PSM4_dict[str(value) + 'End'] = []
            PSM4_list.append([])
        # data parsing
        for device in data.values:
            if device[idx_SPort] in Copper or device[idx_LinkType] in ['Serial', 'Mgmt']:
                idx = idx_Start
                while idx <= idx_End:
                    if pd.isnull(device[idx]):
                        pass
                    else:
                        if device[idx_device][-2:] in ['P1', 'P2', 'T1']:
                            if (idx - idx_Start) % 2 == 0:
                                Cooper_dict['Long_Start'].append(device[idx])
                            else:
                                Cooper_dict['Long_End'].append(device[idx])
                        else:
                            if (idx - idx_Start) % 2 == 0:
                                Cooper_dict['Short_Start'].append(device[idx])
                            else:
                                Cooper_dict['Short_End'].append(device[idx])
                    idx += 1
            if device[idx_SPort] == 'Gi0/0/0':
                idx = idx_Start
                while idx <= idx_End:
                    if pd.isnull(device[idx]):
                        pass
                    else:
                        if (idx - idx_Start) // 2 == 0:
                            if (idx - idx_Start) % 2 == 0:
                                LC_dict['Short_Start'].append(device[idx])
                            else:
                                LC_dict['Short_End'].append(device[idx])
                        else:
                            if (idx - idx_Start) % 2 == 0:
                                LC_dict['Long_Start'].append(device[idx])
                            else:
                                LC_dict['Long_End'].append(device[idx])
                    idx += 1
            if device[idx_device][-2:] == 'M0' and device[idx_SPort][:3] == 'ETH':
                idx = idx_Start
                while idx <= idx_End:
                    if (idx - idx_Start) // 2 == 0:
                        if (idx - idx_Start) % 2 == 0:
                            LC_dict['Short_Start'].append(device[idx])
                        else:
                            LC_dict['Short_End'].append(device[idx])
                    elif (idx - idx_Start) // 2 == 1:
                        if (idx - idx_Start) % 2 == 0:
                            LC_dict['Long_Start'].append(device[idx])
                        else:
                            LC_dict['Long_End'].append(device[idx])
                    else:
                        if (idx - idx_Start) % 2 == 0:
                            MPO['Start'].append(device[idx])
                        else:
                            MPO['End'].append(device[idx][:-1] + "1-4")
                    idx += 1
            if device[idx_device][-2:] == 'T1' and device[idx_SPort][:3] == 'ETH':
                idx = idx_Start
                while idx <= idx_End:
                    if pd.isnull(device[idx]):
                        pass
                    else:
                        if (idx - idx_Start) // 2 == 0:
                            if (idx - idx_Start) % 2 == 0:
                                PSM4_dict_2m['Start'].append(device[idx])
                            else:
                                PSM4_dict_2m['End'].append(device[idx])
                        else:
                            id = (int(device[idx_EndDev][-4:-2]) - 1) % 8
                            if (idx - idx_Start) % 2 == 0:
                                PSM4_dict[str(id) + 'Start'].append(device[idx])
                            else:
                                PSM4_dict[str(id) + 'End'].append(device[idx])
                    idx += 1

        # Cooper_dict['Short_Start'].pop(0)
        # Cooper_dict['Short_End'].pop(0)

        # Merge list
        Cooper_Long = [i + "\n" + j for i, j in zip(Cooper_dict['Long_Start'], Cooper_dict['Long_End'])]
        Copper_Short = [i + "\n" + j for i, j in zip(Cooper_dict['Short_Start'], Cooper_dict['Short_End'])]
        LC_Long = [i + "\n" + j for i, j in zip(LC_dict['Long_Start'], LC_dict['Long_End'])]
        LC_Short = [i + "\n" + j for i, j in zip(LC_dict['Short_Start'], LC_dict['Short_End'])]
        MPO_full = [i + "\n" + j for i, j in zip(MPO['Start'], MPO['End'])]
        PSM4_2m = [i + "\n" + j for i, j in zip(PSM4_dict_2m['Start'], PSM4_dict_2m['End'])]
        for value in range(T1_num):
            PSM4_list[value] = [i + "\n" + j for i, j in
                                zip(PSM4_dict[str(value) + 'Start'], PSM4_dict[str(value) + 'End'])]

        # MOR SIDE
        self.w1.cell(2, 1).value = "MOR SIDE"
        self.w1.cell(2, 1).font = self.b_tile_font

        for name in Cooper_dict.keys():
            self.w1.append(Cooper_dict[name] * 2)

        # Copper Short
        copper_start_l = 3
        self.w1.insert_rows(copper_start_l)  # 3 is start number
        self.w1.cell(copper_start_l, 1).value = "Long Copper [12FT]"
        self.w1.cell(copper_start_l, 2).value = str(len(Cooper_dict['Long_Start'])) + "EA"
        self.w1.cell(copper_start_l, 1).font = self.tile_font

        self.w1.insert_rows(copper_start_l + 3)
        for i in range(len(Cooper_Long) * 2):
            self.w1.cell(copper_start_l + 3, i + 1).value = Cooper_Long[i % len(Cooper_Long)]
            self.w1[copper_start_l + 3][i].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            self.w1[copper_start_l + 3][i].font = Font(size=5.5)

        # Copper Long
        copper_start_s = copper_start_l + 4
        self.w1.insert_rows(copper_start_s)
        self.w1.cell(copper_start_s, 1).value = "Short Copper[8FT]"
        self.w1.cell(copper_start_s, 2).value = str(len(Cooper_dict['Short_Start'])) + "EA"
        self.w1.cell(copper_start_s, 1).font = self.tile_font

        self.w1.insert_rows(copper_start_s + 3)
        for i in range(len(Copper_Short) * 2):
            self.w1.cell(copper_start_s + 3, i + 1).value = Copper_Short[i % len(Copper_Short)]
            self.w1[copper_start_s + 3][i].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            self.w1[copper_start_s + 3][i].font = Font(size=5.5)

        for name in LC_dict.keys():
            if "Short" in name:
                self.w1.append(LC_dict[name] * 2)

        # LC short
        LC_s = copper_start_s + 4
        self.w1.insert_rows(LC_s)
        self.w1.cell(LC_s, 1).value = "SM-LC (2m)"
        self.w1.cell(LC_s, 2).value = str(len(LC_dict['Short_Start'])) + "EA"
        self.w1.cell(LC_s, 1).font = self.tile_font

        self.w1.insert_rows(LC_s + 3)
        for i in range(len(LC_Short) * 2):
            self.w1.cell(LC_s + 3, i + 1).value = LC_Short[i % len(LC_Short)]
            self.w1[LC_s + 3][i].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            self.w1[LC_s + 3][i].font = Font(size=5.5)

        # PSM4 Start
        PSM4_start = LC_s + 4
        self.w1.insert_rows(PSM4_start)
        self.w1.cell(PSM4_start, 1).value = "PSM4 (2m)"
        self.w1.cell(PSM4_start, 2).value = str(len(PSM4_dict_2m['Start'])) + "EA"
        self.w1.cell(PSM4_start, 1).font = self.tile_font
        self.w1.append(PSM4_dict_2m['Start'] * 2)
        self.w1.append(PSM4_dict_2m['End'] * 2)

        self.w1.insert_rows(PSM4_start + 3)
        for i in range(len(PSM4_2m) * 2):
            self.w1.cell(PSM4_start + 3, i + 1).value = PSM4_2m[i % len(PSM4_2m)]
            self.w1[PSM4_start + 3][i].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            self.w1[PSM4_start + 3][i].font = Font(size=5.5)

        # PSM4 2m
        PSM4_T1 = PSM4_start + 4
        self.w1.cell(PSM4_start, 1).value = "PSM4 (2m)"
        self.w1.cell(PSM4_start, 2).value = str(len(PSM4_dict_2m['Start'])) + "EA"
        self.w1.cell(PSM4_start, 1).font = self.tile_font

        # T1 2m
        T1_s = PSM4_T1 + 1
        for i in range(T1_num):
            self.w1.cell(T1_s + (4 * i), 1).value = PSM4_dict_2m['Start'][8 * i][:11]
            self.w1.cell(T1_s + (4 * i), 1).font = self.s_tile_font
            self.w1.append(PSM4_dict_2m['Start'][T1_num * i:T1_num * (i + 1)] * 2)
            self.w1.append(PSM4_dict_2m['End'][T1_num * i:T1_num * (i + 1)] * 2)
            self.w1.append(PSM4_2m[T1_num * i:T1_num * (i + 1)] * 2)
            for cell1 in self.w1[T1_s + (4 * i) + 3]:
                cell1.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell1.font = Font(size=5.5)

        # IDF SIDE
        IDF_start = T1_s + 34
        self.w1.cell(IDF_start, 1).value = "IDF SIDE"
        self.w1.cell(IDF_start, 1).font = self.b_tile_font

        self.w1.cell(IDF_start + 1, 1).value = "SM-LC (11m)"
        self.w1.cell(IDF_start + 1, 2).value = str(len(LC_dict['Long_Start'])) + "EA"
        self.w1.cell(IDF_start + 1, 1).font = self.tile_font

        for name in LC_dict.keys():
            if "Long" in name:
                self.w1.append(LC_dict[name] * 2)
        self.w1.append(LC_Long * 2)
        for cell1 in self.w1[IDF_start + 4]:
            cell1.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell1.font = Font(size=5.5)

        # PSM4 T2
        PSM4_T2 = IDF_start + 5
        self.w1.cell(PSM4_T2, 1).value = "PSM4 "
        self.w1.cell(PSM4_T2, 2).value = "Each T2 " + str(len(PSM4_dict['0Start'])) + "EA"
        self.w1.cell(PSM4_T2, 1).font = self.tile_font

        for i in range(T1_num):
            self.w1.cell(PSM4_T2 + 1 + (4 * i), 1).value = str(i + 1) + "&" + str(i + 9) + "T2"
            self.w1.cell(PSM4_T2 + 1 + (4 * i), 1).font = self.s_tile_font
            self.w1.append(PSM4_dict[str(i) + 'Start'] * 2)
            self.w1.append(PSM4_dict[str(i) + 'End'] * 2)
            self.w1.append(PSM4_list[i] * 2)
            for cell1 in self.w1[PSM4_T2 + 4 + (4 * i)]:
                cell1.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell1.font = Font(size=5.5)

        MPO_row = PSM4_T2 + 35

        self.w1.cell(MPO_row, 1).value = "MPO "
        self.w1.cell(MPO_row, 2).value = str(len(MPO['Start'])) + "EA"
        self.w1.cell(MPO_row, 3).value = "Option"
        self.w1.cell(MPO_row, 1).font = self.tile_font

        for name in MPO.keys():
            self.w1.append(MPO[name] * 2)
        for i in range(len(MPO_full) * 2):
            self.w1.cell(MPO_row + 3, i + 1).value = MPO_full[i % len(MPO_full)]
            self.w1[MPO_row + 3][i].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            self.w1[MPO_row + 3][i].font = Font(size=5.5)

        for col in range(1, len(PSM4_2m) + 1):
            self.w1.column_dimensions[get_column_letter(col)].width = 24
            self.w2.column_dimensions[get_column_letter(col)].width = 24

        self.wb.save(load_file_name+"_label_filter.xlsx")
        QMessageBox.about(self, 'Alert', 'Completed!')
        QCoreApplication.instance().quit()


app = QApplication(sys.argv)
ex = MyApp()
sys.exit(app.exec_())
