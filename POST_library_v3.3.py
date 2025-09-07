import os
import sys
import pandas as pd
import time
from openpyxl import load_workbook

from PyQt5.QtWidgets import QApplication, QWidget, QPushButton, QHBoxLayout, QVBoxLayout, QComboBox, QLabel, QLineEdit, QTextEdit, QMessageBox
from PyQt5.QtCore import QCoreApplication

now_path = os.getcwd()

class MyApp(QWidget):

    def __init__(self):
        super().__init__()
        # Default value
        # self.meta_data = pd.read_excel(now_path + '\POST_CODE.xlsx', sheet_name=None, na_values="TT")

        self.alias = QLineEdit('name', self)
        self.file_name = QLineEdit('file', self)
        if os.path.isfile(now_path + '\\address_file.txt'):
            with open(now_path + '\\address_file.txt', 'r') as file:
                self.file_name.setText(file.readline()[:-1])
                self.alias.setText(file.readline())

        self.folder = QLineEdit(now_path, self)
        self.meta_data = []
        self.m_name = QLabel('None', self)
        self.post = QLineEdit('AA', self)
        self.post.setMaxLength(2)
        self.S_des = QTextEdit('According to the manufacturer', self)
        self.D_des = QTextEdit('According to the DCT experience', self)
        self.note = QLineEdit('note', self)
        self.today = time.strftime('%Y-%m-%d', time.localtime(time.time()))

        self.ex_row = 0
        self.manu, self.relate_task, self.note_his, self.a_name, self.date, self.DC_code = 3, 4, 5, 6, 7, 8

        self.initUI()

    def initUI(self):

        cb = QComboBox(self)
        for item in ['Wiwynn', 'Quanta', "Ingrasys", "Lenovo", "DELL", "HPE"]:
            cb.addItem(item)
        cb.activated[str].connect(self.setting_name)

        F_Button = QPushButton('Go!', self)
        M_Button = QPushButton('Search!', self)
        S_Button = QPushButton('save', self)
        cancelButton = QPushButton('Exit', self)

        F_Button.clicked.connect(self.setting_address)
        cancelButton.clicked.connect(QCoreApplication.instance().quit)
        M_Button.clicked.connect(self.search)
        M_Button.clicked.connect(self.DCT_search)
        S_Button.clicked.connect(self.input_data)

        title0 = QHBoxLayout()
        title0.addWidget(QLabel('Path'))
        title0.addWidget(self.folder)
        title0.addWidget(QLabel('file_name'))
        title0.addWidget(self.file_name)

        title = QHBoxLayout()
        title.addWidget(QLabel('Alias'))
        title.addWidget(self.alias)
        title.addWidget(F_Button)

        hbox0 = QHBoxLayout()
        hbox0.addWidget(QLabel('Manufacturer'))
        hbox0.addWidget(cb)
        hbox0.addWidget(self.m_name)
        hbox0.addStretch(1)

        hbox1 = QHBoxLayout()
        hbox1.addWidget(QLabel('POST CODE'))
        hbox1.addWidget(self.post)
        hbox1.addWidget(M_Button)
        hbox1.addStretch(1)

        hbox2 = QHBoxLayout()
        hbox2.addWidget(QLabel('Specification'))
        hbox2.addStretch(1)

        hbox3 = QHBoxLayout()
        hbox3.addWidget(self.S_des)
        hbox3.addStretch(1)

        hbox4 = QHBoxLayout()
        hbox4.addWidget(QLabel('DCT Comments'))
        hbox4.addStretch(1)

        hbox5 = QHBoxLayout()
        hbox5.addWidget(self.D_des)
        hbox5.addStretch(1)

        hbox6 = QHBoxLayout()
        hbox6.addWidget(QLabel('Note'))
        hbox6.addWidget(self.note)
        hbox6.addWidget(S_Button)
        hbox6.addStretch(1)

        hbox7 = QHBoxLayout()
        hbox7.addWidget(cancelButton)
        hbox7.addStretch(1)

        vbox = QVBoxLayout()
        vbox.addStretch(1)
        vbox.addLayout(title0)
        vbox.addStretch(1)
        vbox.addLayout(title)
        vbox.addStretch(1)
        vbox.addLayout(hbox0)
        vbox.addStretch(1)
        vbox.addLayout(hbox1)
        vbox.addStretch(1)
        vbox.addLayout(hbox2)
        vbox.addStretch(1)
        vbox.addLayout(hbox3)
        vbox.addStretch(1)
        vbox.addLayout(hbox4)
        vbox.addStretch(1)
        vbox.addLayout(hbox5)
        vbox.addStretch(1)
        vbox.addLayout(hbox6)
        vbox.addStretch(1)
        vbox.addLayout(hbox7)
        vbox.addStretch(1)

        self.setLayout(vbox)

        self.setWindowTitle('POST Library')
        self.setGeometry(100, 100, 250, 600)
        self.show()

    def setting_name(self, text):
        self.m_name.setText(text)

    def setting_address(self):
        self.address = self.folder.text() + '\\' + self.file_name.text()
        self.meta_data = pd.read_excel(self.address, sheet_name=None, na_values="TT")
        with open(now_path + '\\address_file.txt', 'w') as file:
            file.write(self.file_name.text()+"\n")
            file.write(self.alias.text())

    def input_data(self):
        self.work_book = load_workbook(self.address, data_only=True)
        load_ws = self.work_book['DCT']

        load_ws.insert_rows(self.ex_row)
        load_ws.cell(self.ex_row, 1, value=self.post.text())
        load_ws.cell(self.ex_row, self.manu, value=self.m_name.text())
        load_ws.cell(self.ex_row, self.note_his, value=self.note.text())
        load_ws.cell(self.ex_row, self.a_name, value=self.alias.text())
        load_ws.cell(self.ex_row, self.date, value=self.today)
        load_ws.cell(self.ex_row, self.DC_code, value='PUS04')

        self.work_book.save(self.address)
        self.meta_data = pd.read_excel(self.address, sheet_name=None, na_values="TT")

    def search(self):
        if self.m_name.text() == 'None' or self.meta_data == []:
            QMessageBox.about(self, 'Alert', 'Please select manufacturer')
        else:
            sheet = self.m_name.text()
            data = self.meta_data[sheet]
            data = data.fillna('')
            find_code = self.post.text().upper()
            description = ''
            for value in data.values:
                if value[0] == find_code:
                    description += value[1]
                    description += '\n'
            if description == '':
                self.S_des.setText('No found')
            else:
                self.S_des.setText(description)

    def DCT_search(self):
        if self.m_name.text() == 'None' or self.meta_data == []:
            pass
        else:
            data = self.meta_data['DCT']
            data = data.fillna('')
            find_code = self.post.text().upper()
            note = ''
            for idx, value in enumerate(data.values):
                if value[0] == find_code: # BIOS code self.name
                    self.ex_row = idx + 3
                    if value[self.date] != '':
                        # note += ('GDCO : ' + value[self.relate_task-1][-10:] + " | ")
                        note += ('date : ' + str(value[self.date-1])[:10] + " | ")
                        note += ('DC : ' + value[self.DC_code-1])
                        note += '\n'
                        note += ('note : ' + value[self.note_his-1])
                        note += '\n'
                    else:
                        pass
            if note == '':
                self.D_des.setText('No found')
            else:
                self.D_des.setText(note)


app = QApplication(sys.argv)
ex = MyApp()
sys.exit(app.exec_())