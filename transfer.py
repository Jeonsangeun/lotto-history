import sys

import openpyxl as excel
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill, colors, numbers
import pandas as pd
import os

wb = excel.Workbook()
wb.active.title = 'label'
wb.create_sheet('template1', 1)

w1 = wb['label']
w2 = wb['template1']

save_file_name = "example.xlsx"
file_list = os.listdir()

# writing
style_center = Alignment(horizontal='center', vertical='center')
title_font = Font(size=20, bold=True, color='ffffff')
tile_font = Font(size=15, bold=True, color='000000')
subject_bg = PatternFill('solid', fgColor='4BACC6')
AOC_bg = PatternFill('solid', fgColor='ff0000')
copper_bg = PatternFill('solid', fgColor='0000FF')

w1.merge_cells('A1:G1')
w1['A1'] = "Label print maker"
w1['A1'].font = title_font
w1['A1'].fill = subject_bg
w1['A1'].alignment = style_center

w2.merge_cells('A1:G1')
w2['A1'] = "Label template"
w2['A1'].font = title_font
w2['A1'].fill = subject_bg
w2['A1'].alignment = style_center

def PRD(load_file_name):
    # import NDT data
    # load_file_name = "PRD.xlsx"
    # load_file_name = file_list[0]
    sheet_name = "Sheet1"
    default = 0
    data = pd.read_excel(load_file_name, sheet_name=default, na_values='=')
    max_length = 0
    dump = 4
    rack_dump = []
    print(data.keys())

    # Observe all Start point
    AOC_dict_start, total_AOC_s = dict(), []
    AOC_dict_end, total_AOC_e = dict(), []
    COP_dict_start, total_COP_s = dict(), []
    COP_dict_end, total_COP_e = dict(), []
    for idx, value in enumerate(data.values):
        # print(idx, " : ", value[3][12:15])
        AOC_dict_start[value[3][:5]] = []
        AOC_dict_end[value[3][:5]] = []
        COP_dict_start[value[3][:5]] = []
        COP_dict_end[value[3][:5]] = []

    # insert table
    for idx, value in enumerate(data.values):
        if value[1][:3] == 'ETH':
            AOC_dict_start[value[3][:5]].append(value[3])
            AOC_dict_end[value[3][:5]].append(value[4])
        else:
            COP_dict_start[value[3][:5]].append(value[3])
            COP_dict_end[value[3][:5]].append(value[4])

    Key = list(AOC_dict_start.keys())
    temp_1, temp_2, temp_3, temp_4, temp_reack = [], [], [], [], []
    # write in w1
    for id, text in enumerate(Key):
        w1.cell(7 * id + 2, 1).value = text
        w1.cell(7 * id + 2, 1).font = tile_font
        w1.cell(7 * id + 3, 1).value = "AOC cable"
        w1.cell(7 * id + 3, 1).fill = AOC_bg
        w1.append(AOC_dict_start[text] * 2)
        w1.append(AOC_dict_end[text] * 2)
        w1.cell(7 * id + 6, 1).value = "Copper cable"
        w1.cell(7 * id + 6, 1).fill = copper_bg
        w1.append(COP_dict_start[text] * 2)
        w1.append(COP_dict_end[text] * 2)

        if id % dump == 0 and id != 0:
            total_AOC_s.append(temp_1)
            total_AOC_e.append(temp_2)
            total_COP_s.append(temp_3)
            total_COP_e.append(temp_4)
            rack_dump.append(temp_reack)
            temp_1, temp_2, temp_3, temp_4, temp_reack = [], [], [], [], []

        temp_1 += AOC_dict_start[text] * 2
        temp_2 += AOC_dict_end[text] * 2
        temp_3 += COP_dict_start[text] * 2
        temp_4 += COP_dict_end[text] * 2
        temp_reack.append(text)

        if max_length < len(AOC_dict_start[text] * 2):
            max_length = len(AOC_dict_start[text] * 2)

    for id, label in enumerate(rack_dump):
        w2.cell(8 * id + 2, 2).value = "Cop_label" + str(2*len(total_COP_s[id])) + "EA"
        w2.cell(8 * id + 2, 1).value = "AOC_label" + str(2*len(total_AOC_s[id])) + "EA"
        w2.append(label)
        w2.cell(8 * id + 4, 1).value = "AOC cable"
        w2.cell(8 * id + 4, 1).fill = AOC_bg
        w2.append(total_AOC_s[id])
        w2.append(total_AOC_e[id])
        w2.cell(8 * id + 7, 1).value = "Copper cable"
        w2.cell(8 * id + 7, 1).fill = copper_bg
        w2.append(total_COP_s[id])
        w2.append(total_COP_e[id])

    for col in range(1, max_length + 1):
        w1.column_dimensions[get_column_letter(col)].width = 22
        w2.column_dimensions[get_column_letter(col)].width = 22

    for col in range(1, len(total_AOC_s) + 1):
        w2.column_dimensions[get_column_letter(col)].width = 22
        w2.column_dimensions[get_column_letter(col)].width = 22

def main():
    file_list = os.listdir()
    print(file_list)
    PRD("PRD.xlsx")
    wb.save(save_file_name)

if __name__ == '__main__':
    main()