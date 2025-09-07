import sys

import openpyxl as excel
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill, colors, numbers
import pandas as pd
import os

wb = excel.Workbook()
wb.active.title = 'label'
wb.create_sheet('print_model1', 1)

w1 = wb['label']
w2 = wb['print_model1']

save_file_name = "label_filter.xlsx"
file_list = os.listdir()

# writing
style_center = Alignment(horizontal='center', vertical='center')
title_font = Font(size=20, bold=True, color='ffffff')
b_tile_font = Font(size=20, bold=True, color='000000')
tile_font = Font(size=15, bold=True, color='000000')
s_tile_font = Font(size=10, bold=True, color='000000')
subject_bg = PatternFill('solid', fgColor='4BACC6')
AOC_bg = PatternFill('solid', fgColor='ff0000')
copper_bg = PatternFill('solid', fgColor='0000FF')

w1.merge_cells('A1:G1')
w1['A1'] = "Label print maker"
w1['A1'].font = title_font
w1['A1'].fill = subject_bg
w1['A1'].alignment = style_center

w2.merge_cells('A1:G1')
w2['A1'] = "Label template"
w2['A1'].font = title_font
w2['A1'].fill = subject_bg
w2['A1'].alignment = style_center

def PRD(load_file_name):
    # import NDT data
    # load_file_name = "PRD.xlsx"
    # load_file_name = file_list[0]
    sheet_name = "Sheet1"
    default = 0
    data = pd.read_excel(load_file_name, sheet_name=default, na_values='=')
    max_length = 0
    dump = 4
    rack_dump = []

    # Observe all Start point
    AOC_dict_start, total_AOC_s = dict(), []
    AOC_dict_end, total_AOC_e = dict(), []
    COP_dict_start, total_COP_s = dict(), []
    COP_dict_end, total_COP_e = dict(), []
    for idx, value in enumerate(data.values):
        # print(idx, " : ", value[3][12:15])
        AOC_dict_start[value[3][:5]] = []
        AOC_dict_end[value[3][:5]] = []
        COP_dict_start[value[3][:5]] = []
        COP_dict_end[value[3][:5]] = []

    # insert table
    for idx, value in enumerate(data.values):
        if value[1][:3] == 'ETH':
            AOC_dict_start[value[3][:5]].append(value[3])
            AOC_dict_end[value[3][:5]].append(value[4])
        else:
            COP_dict_start[value[3][:5]].append(value[3])
            COP_dict_end[value[3][:5]].append(value[4])

    Key = list(AOC_dict_start.keys())
    temp_1, temp_2, temp_3, temp_4, temp_reack = [], [], [], [], []
    # write in w1
    for id, text in enumerate(Key):
        w1.cell(7 * id + 2, 1).value = text
        w1.cell(7 * id + 2, 1).font = tile_font
        w1.cell(7 * id + 3, 1).value = "AOC cable"
        w1.cell(7 * id + 3, 1).fill = AOC_bg
        w1.append(AOC_dict_start[text] * 2)
        w1.append(AOC_dict_end[text] * 2)
        w1.cell(7 * id + 6, 1).value = "Copper cable"
        w1.cell(7 * id + 6, 1).fill = copper_bg
        w1.append(COP_dict_start[text] * 2)
        w1.append(COP_dict_end[text] * 2)

        if id % dump == 0 and id != 0:
            total_AOC_s.append(temp_1)
            total_AOC_e.append(temp_2)
            total_COP_s.append(temp_3)
            total_COP_e.append(temp_4)
            rack_dump.append(temp_reack)
            temp_1, temp_2, temp_3, temp_4, temp_reack = [], [], [], [], []

        temp_1 += AOC_dict_start[text] * 2
        temp_2 += AOC_dict_end[text] * 2
        temp_3 += COP_dict_start[text] * 2
        temp_4 += COP_dict_end[text] * 2
        temp_reack.append(text)

        if max_length < len(AOC_dict_start[text] * 2):
            max_length = len(AOC_dict_start[text] * 2)

    for id, label in enumerate(rack_dump):
        w2.cell(8 * id + 2, 2).value = "Cop_label : " + str(2*len(total_COP_s[id])) + "EA"
        w2.cell(8 * id + 2, 1).value = "AOC_label : " + str(2*len(total_AOC_s[id])) + "EA"
        w2.append(label)
        w2.cell(8 * id + 4, 1).value = "AOC cable"
        w2.cell(8 * id + 4, 1).fill = AOC_bg
        w2.append(total_AOC_s[id])
        w2.append(total_AOC_e[id])
        w2.cell(8 * id + 7, 1).value = "Copper cable"
        w2.cell(8 * id + 7, 1).fill = copper_bg
        w2.append(total_COP_s[id])
        w2.append(total_COP_e[id])

    for col in range(1, max_length + 1):
        w1.column_dimensions[get_column_letter(col)].width = 22
        w2.column_dimensions[get_column_letter(col)].width = 22

    for col in range(1, max_length * dump + 1):
        w2.column_dimensions[get_column_letter(col)].width = 22
        w2.column_dimensions[get_column_letter(col)].width = 22


def MOR(load_file_name):
    # import NDT data
    # load_file_name = "PRD.xlsx"
    # load_file_name = file_list[0]
    sheet_name = "Sheet1"
    default = 0
    data = pd.read_excel(load_file_name, sheet_name=default)
    idx_device, idx_SPort, idx_Start, idx_End = 0, 1, 4, 9
    Copper = ['Gi0/0/1', 'Gi0/0/2', 'CONS', 'MGMT', 'COM6', 'NIC']
    Cooper_dict = {'Long_Start': [], 'Long_End': [], 'Short_Start': [], 'Short_End': []}
    LC_dict = {'Long_Start': [], 'Long_End': [], 'Short_Start': [], 'Short_End': []}
    MPO = {'Start': [], 'End': []}
    PSM4_dict_2m = {'Start': [], 'End': []}
    PSM4_dict = {}
    T1_num = 8
    for value in range(T1_num):
        PSM4_dict[str(value) + 'Start'] = []
        PSM4_dict[str(value) + 'End'] = []
    T1_flag = 0

    for device in data.values:
        if device[idx_SPort] in Copper:
            idx = idx_Start
            while idx <= idx_End:
                if pd.isnull(device[idx]):
                    pass
                else:
                    if device[idx_device][-2:] in ['P1', 'P2', 'T1']:
                        if (idx - idx_Start) % 2 == 0:
                            Cooper_dict['Long_Start'].append(device[idx])
                        else:
                            Cooper_dict['Long_End'].append(device[idx])
                    else:
                        if (idx - idx_Start) % 2 == 0:
                            Cooper_dict['Short_Start'].append(device[idx])
                        else:
                            Cooper_dict['Short_End'].append(device[idx])
                idx += 1
        if device[idx_SPort] == 'Gi0/0/0':
            idx = idx_Start
            while idx <= idx_End:
                if pd.isnull(device[idx]):
                    pass
                else:
                    if (idx - idx_Start) // 2 == 0:
                        if (idx - idx_Start) % 2 == 0:
                            LC_dict['Short_Start'].append(device[idx])
                        else:
                            LC_dict['Short_End'].append(device[idx])
                    else:
                        if (idx - idx_Start) % 2 == 0:
                            LC_dict['Long_Start'].append(device[idx])
                        else:
                            LC_dict['Long_End'].append(device[idx])
                idx += 1
        if device[idx_device][-2:] == 'M0' and device[idx_SPort][:3] == 'ETH':
            idx = idx_Start
            while idx <= idx_End:
                if (idx - idx_Start) // 2 == 0:
                    if (idx - idx_Start) % 2 == 0:
                        LC_dict['Short_Start'].append(device[idx])
                    else:
                        LC_dict['Short_End'].append(device[idx])
                elif (idx - idx_Start) // 2 == 1:
                    if (idx - idx_Start) % 2 == 0:
                        LC_dict['Long_Start'].append(device[idx])
                    else:
                        LC_dict['Long_End'].append(device[idx])
                else:
                    if (idx - idx_Start) % 2 == 0:
                        MPO['Start'].append(device[idx])
                    else:
                        MPO['End'].append(device[idx])
                idx += 1
        if device[idx_device][-2:] == 'T1' and device[idx_SPort][:3] == 'ETH':
            idx = idx_Start
            while idx <= idx_End:
                if pd.isnull(device[idx]):
                    pass
                else:
                    if (idx - idx_Start) // 2 == 0:
                        if (idx - idx_Start) % 2 == 0:
                            PSM4_dict_2m['Start'].append(device[idx])
                        else:
                            PSM4_dict_2m['End'].append(device[idx])
                    else:
                        id = T1_flag % T1_num
                        if (idx - idx_Start) % 2 == 0:
                            PSM4_dict[str(id) + 'Start'].append(device[idx])
                        else:
                            PSM4_dict[str(id) + 'End'].append(device[idx])
                idx += 1
            T1_flag += 1
    Cooper_dict['Short_Start'].pop(0)
    Cooper_dict['Short_End'].pop(0)

    # MOR SIDE
    w1.cell(2, 1).value = "MOR SIDE"
    w1.cell(2, 1).font = b_tile_font

    for name in Cooper_dict.keys():
        w1.append(Cooper_dict[name] * 2)

    w1.insert_rows(3)
    w1.cell(3, 1).value = "Long Copper [20m]"
    w1.cell(3, 2).value = str(len(Cooper_dict['Long_Start'])) + "EA"
    w1.cell(3, 1).font = tile_font

    w1.insert_rows(6)
    w1.cell(6, 1).value = "Short Copper[7m]"
    w1.cell(6, 2).value = str(len(Cooper_dict['Short_Start'])) + "EA"
    w1.cell(6, 1).font = tile_font

    w1.cell(9, 1).value = "SM-LC (2m)"
    w1.cell(9, 2).value = str(len(LC_dict['Short_Start'])) + "EA"
    w1.cell(9, 1).font = tile_font

    for name in LC_dict.keys():
        if "Short" in name:
            w1.append(LC_dict[name] * 2)

    w1.cell(12, 1).value = "PSM4 (2m)"
    w1.cell(12, 2).value = str(len(PSM4_dict_2m['Start'])) + "EA"
    w1.cell(12, 1).font = tile_font

    for i in range(T1_num):
        w1.cell(13+(3*i), 1).value = PSM4_dict_2m['Start'][8*i][:11]
        w1.cell(13+(3*i), 1).font = s_tile_font
        w1.append(PSM4_dict_2m['Start'][8*i:8*(i+1)] * 2)
        w1.append(PSM4_dict_2m['End'][8 * i:8 * (i + 1)] * 2)

     # MOR SIDE
    w1.cell(40, 1).value = "IDF SIDE"
    w1.cell(40, 1).font = b_tile_font

    w1.cell(41, 1).value = "SM-LC (9m)"
    w1.cell(41, 2).value = str(len(LC_dict['Long_Start'])) + "EA"
    w1.cell(41, 1).font = tile_font

    for name in LC_dict.keys():
        if "Long" in name:
            w1.append(LC_dict[name] * 2)

    w1.cell(44, 1).value = "PSM4 "
    w1.cell(44, 2).value = "Each T2 " + str(len(PSM4_dict['0Start'])) + "EA"
    w1.cell(44, 1).font = tile_font

    for i in range(T1_num):
        w1.cell(45 + (3 * i), 1).value = str(PSM4_dict[str(i)+'Start'][0][:2]) + " | " + str(i+1) + "&" + str(i+9) + "T2"
        w1.cell(45 + (3 * i), 1).font = s_tile_font
        w1.append(PSM4_dict[str(i)+'Start'] * 2)
        w1.append(PSM4_dict[str(i)+'End'] * 2)

    for col in range(1, len(PSM4_dict_2m['Start']) + 1):
        w1.column_dimensions[get_column_letter(col)].width = 24
        w2.column_dimensions[get_column_letter(col)].width = 24

def main():
    file_list = os.listdir()
    if "MOR.xlsx" in file_list:
        MOR("MOR.xlsx")
    elif "PRD.xlsx" in file_list:
        PRD("PRD.xlsx")

    wb.save(save_file_name)

if __name__ == '__main__':
    main()