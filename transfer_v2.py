import sys
from PyQt5.QtWidgets import QApplication, QWidget, QPushButton, QHBoxLayout, QVBoxLayout, QComboBox, QLabel
from PyQt5.QtCore import QCoreApplication

import openpyxl as excel
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill
import pandas as pd
import os

class MyApp(QWidget):

    def __init__(self):
        super().__init__()
        self.wb = excel.Workbook()
        self.wb.active.title = 'label'
        self.wb.create_sheet('template1', 1)
        self.w1 = self.wb['label']
        self.w2 = self.wb['template1']

        # writing
        self.style_center = Alignment(horizontal='center', vertical='center')
        self.title_font = Font(size=20, bold=True, color='ffffff')
        self.tile_font = Font(size=15, bold=True, color='000000')
        self.subject_bg = PatternFill('solid', fgColor='4BACC6')
        self.AOC_bg = PatternFill('solid', fgColor='ff0000')
        self.copper_bg = PatternFill('solid', fgColor='0000FF')

        self.w1.merge_cells('A1:G1')
        self.w1['A1'] = "Label print maker"
        self.w1['A1'].font = self.title_font
        self.w1['A1'].fill = self.subject_bg
        self.w1['A1'].alignment = self.style_center

        self.w2.merge_cells('A1:G1')
        self.w2['A1'] = "Label template"
        self.w2['A1'].font = self.title_font
        self.w2['A1'].fill = self.subject_bg
        self.w2['A1'].alignment = self.style_center

        self.initUI()

    def initUI(self):

        okButton = QPushButton('OK', self)
        cancelButton = QPushButton('Cancel', self)

        self.name = QLabel('Option1', self)
        self.name.move(50, 150)

        file_list = os.listdir()
        cb = QComboBox(self)
        for item in file_list:
            cb.addItem(item)
        cb.move(50, 50)

        cb.activated[str].connect(self.setting_name)

        okButton.clicked.connect(self.make_PRD)
        cancelButton.clicked.connect(QCoreApplication.instance().quit)

        hbox = QHBoxLayout()
        hbox.addStretch(1)
        hbox.addWidget(okButton)
        hbox.addWidget(cancelButton)
        hbox.addStretch(1)

        hbox1 = QHBoxLayout()
        hbox1.addStretch(1)
        hbox1.addWidget(cb)
        hbox1.addWidget(self.name)
        hbox1.addStretch(1)

        vbox = QVBoxLayout()
        vbox.addStretch(1)
        vbox.addLayout(hbox1)
        vbox.addStretch(1)
        vbox.addLayout(hbox)
        vbox.addStretch(1)

        self.setLayout(vbox)

        self.setWindowTitle('Box Layout')
        self.setGeometry(300, 300, 300, 200)
        self.show()

    def setting_name(self, text):
        self.name.setText(text)

    def make_PRD(self):
        default = 0
        # load_file_name = "PRD.xlsx"
        load_file_name = self.name.text()
        data = pd.read_excel(load_file_name, sheet_name=default, na_values='=')
        max_length = 0
        dump = 4
        rack_dump = []

        # Observe all Start point
        AOC_dict_start, total_AOC_s = dict(), []
        AOC_dict_end, total_AOC_e = dict(), []
        COP_dict_start, total_COP_s = dict(), []
        COP_dict_end, total_COP_e = dict(), []
        for idx, value in enumerate(data.values):
            # print(idx, " : ", value[3][12:15])
            AOC_dict_start[value[3][:5]] = []
            AOC_dict_end[value[3][:5]] = []
            COP_dict_start[value[3][:5]] = []
            COP_dict_end[value[3][:5]] = []

        # insert table
        for idx, value in enumerate(data.values):
            if value[1][:3] == 'ETH':
                AOC_dict_start[value[3][:5]].append(value[3])
                AOC_dict_end[value[3][:5]].append(value[4])
            else:
                COP_dict_start[value[3][:5]].append(value[3])
                COP_dict_end[value[3][:5]].append(value[4])

        Key = list(AOC_dict_start.keys())
        temp_1, temp_2, temp_3, temp_4, temp_reack = [], [], [], [], []
        # write in w1
        for id, text in enumerate(Key):
            self.w1.cell(7 * id + 2, 1).value = text
            self.w1.cell(7 * id + 2, 1).font = self.tile_font
            self.w1.cell(7 * id + 3, 1).value = "AOC cable"
            self.w1.cell(7 * id + 3, 1).fill = self.AOC_bg
            self.w1.append(AOC_dict_start[text] * 2)
            self.w1.append(AOC_dict_end[text] * 2)
            self.w1.cell(7 * id + 6, 1).value = "Copper cable"
            self.w1.cell(7 * id + 6, 1).fill = self.copper_bg
            self.w1.append(COP_dict_start[text] * 2)
            self.w1.append(COP_dict_end[text] * 2)

            if id % dump == 0 and id != 0:
                total_AOC_s.append(temp_1)
                total_AOC_e.append(temp_2)
                total_COP_s.append(temp_3)
                total_COP_e.append(temp_4)
                rack_dump.append(temp_reack)
                temp_1, temp_2, temp_3, temp_4, temp_reack = [], [], [], [], []

            temp_1 += AOC_dict_start[text] * 2
            temp_2 += AOC_dict_end[text] * 2
            temp_3 += COP_dict_start[text] * 2
            temp_4 += COP_dict_end[text] * 2
            temp_reack.append(text)

            if max_length < len(AOC_dict_start[text] * 2):
                max_length = len(AOC_dict_start[text] * 2)

        for id, label in enumerate(rack_dump):
            self.w2.cell(8 * id + 2, 2).value = "Cop_label" + str(2 * len(total_COP_s[id])) + "EA"
            self.w2.cell(8 * id + 2, 1).value = "AOC_label" + str(2 * len(total_AOC_s[id])) + "EA"
            self.w2.append(label)
            self.w2.cell(8 * id + 4, 1).value = "AOC cable"
            self.w2.cell(8 * id + 4, 1).fill = self.AOC_bg
            self.w2.append(total_AOC_s[id])
            self.w2.append(total_AOC_e[id])
            self.w2.cell(8 * id + 7, 1).value = "Copper cable"
            self.w2.cell(8 * id + 7, 1).fill = self.copper_bg
            self.w2.append(total_COP_s[id])
            self.w2.append(total_COP_e[id])

        for col in range(1, max_length + 1):
            self.w1.column_dimensions[get_column_letter(col)].width = 22
            self.w2.column_dimensions[get_column_letter(col)].width = 22

        for col in range(1, len(total_AOC_s) + 1):
            self.w2.column_dimensions[get_column_letter(col)].width = 22
            self.w2.column_dimensions[get_column_letter(col)].width = 22

        self.wb.save("label_filter.xlsx")
        QCoreApplication.instance().quit()

def main():
    app = QApplication(sys.argv)
    ex = MyApp()
    sys.exit(app.exec_())

if __name__ == '__main__':
    main()